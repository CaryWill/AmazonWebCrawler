from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.firefox.options import Options
from selenium.webdriver import Firefox
from selenium.webdriver.common.keys import Keys
import lxml.html
#import pymongo
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import string
#Time
from datetime import datetime, date, time

targetProductNameMatching = 'Maevis Bed Waterproof Mattress'
#browser = webdriver.Firefox()
#Headless firefox config
options = Options()
#options.add_argument('-headless')
browser = Firefox(executable_path='geckodriver', firefox_options=options)
wait = WebDriverWait(browser, 10)
browser.set_window_size(1400, 900)

#Excel part
wb = Workbook()
products = []

def search(keyword,pageNumber,sheetNumber,worksheet,myProductIDForMatcting):
    print('正在搜索')
    try:
        #美亚
        browser.get('https://www.amazon.com/')
        input = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#twotabsearchtextbox')))
        submit = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '.nav-search-submit > input:nth-child(2)')))
        input.send_keys(keyword)
        submit.click()
        get_products(keyword,pageNumber,sheetNumber,worksheet,myProductIDForMatcting)
    except TimeoutException:
        return search(keyword,pageNumber,sheetNumber,worksheet,myProductIDForMatcting)


def next_page(keyword,pageNumber,sheetNumber,worksheet,myProductIDForMatcting):
    print('正在翻页', pageNumber)
    try:
        wait.until(EC.text_to_be_present_in_element(
            (By.CSS_SELECTOR, '#pagnNextString'), 'Next Page'))
        submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#pagnNextString')))
        submit.click()
        wait.until(EC.text_to_be_present_in_element(
            (By.CSS_SELECTOR, '.pagnCur'), str(pageNumber)))
        get_products(keyword,pageNumber,sheetNumber,worksheet,myProductIDForMatcting)
    except TimeoutException:
        next_page(keyword,pageNumber,sheetNumber,worksheet,myProductIDForMatcting)


def get_products(keyword,pageNumber,sheetNumber,worksheet,myProductIDForMatcting):
    try:
        #切换为3列模式，不然数据是列模式的。
        #condition1:有9宫格按钮
        #threeColoumMode = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.s-grid-layout-picker')))
        #condition2:只有四格按钮
        #threeColoumMode = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'span.s-layout-toggle-picker:nth-child(3)'))) 
        #threeColoumMode.click()
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#s-results-list-atf')))
        html = browser.page_source
        soup = BeautifulSoup(html, 'lxml')
        #result_0是第一个
        content = soup.find_all(attrs={"id": re.compile(r'result_\d+')})
        print("how many result were found:",len(content))
        #获取产品链接
        #利用图片的父node来提取Tag a里的链接
        #而不是所有的result图片都可以用('img', attrs={'class':'s-access-image cfMarker'})提取到
        #像亚马逊的那种Look for mattress的result就没有s-access-image cfMarker 这个class的img Tag就没有
        #但是每一个result都要img这个tag而且img这个tag的父tag就是包含链接的那个tag
        urls = []
        """
        #BUG-Excel里产品少了一个 最后一个产品没有
        #how many result were found: 42
        #url: 41
        
        productDivs = soup.findAll('img', attrs={'class':'s-access-image cfMarker'})
        #print("links:",len(productDivs))
        for div in productDivs:
            urls.append(div.parent['href'])
        print("url:",len(urls))"""
        #因为result_ container第一个a tag就是那个包含link的所以可以从这里入手
        for div in content:
            urls.append(div.img.parent['href'])
        print("urls:",len(urls))
        #如果有搜索结果
        if len(content)!=0:
            #这里除了问题 url和content数量不一致 导致url少的先结束for loop
            for (item,index,url) in zip(content,range(1,9999),urls):
                product = {
                    #是那种非产品的缺少s-access-title的，就默认给个title
                    #此处的get_text不会引 'NoneType' object has no attribute 'get_text'
                    'title': item.find(class_='s-access-title').get_text() if item.find(class_='s-access-title') else "Amazon recommendation", 
                    'index': index,#在一页里的顺位序号，每一页都会变
                    #'rank': getRank(pageNumber,index),#就算有那种AD也是准的，不影响
                    #'image':item.find(class_='s-access-image cfMarker').get('src')
                }
                #Generate Rank attr for product
                turnProductIndexToRank(product,pageNumber)
                product['link'] = 'https://www.amazon.com'+url if ('Sponsored' in product['title']) else url
                #print(product['link'])
                products.append(product)
                print(index," product is processed.")
        else:
            print("No product were found!")
        #print("how many prodcuts:",len(products))
    except Exception as err:
        print(err)

def saveToExcel(products,keyword):
    try:
        for product in products:
            productURL = product['link']
            productDetail = getProductDetail(productURL)
            print("title:",product['title'])
            print('star rank:',productDetail['starRank'])
            print('review count:',productDetail['reviewCount'])
            print('combined:',productDetail['combine_all_size_priceString']) 
            print('image link:',productDetail['imageLink'])
            print('product link:',product['link']) 
            productDataNeedToSave = [product['title'],productDetail['starRank'],productDetail['reviewCount'],productDetail['combine_all_size_priceString'],productDetail['imageLink'],product['link']]
            wb[keyword].append(productDataNeedToSave)
        wb.save("sample.xlsx")
    except Exception as err:
        print('Save failed:', err)
        wb.save("sample.xlsx") 

def getProductDetail(productURL):
    try:
        #BUG:速度是不是会很慢如果用了很多的find_all() 
        browser.get(productURL) 
        html = browser.page_source
        soup = BeautifulSoup(html, 'lxml')
        #product info
        sizes = getAll_Size_PriceForEachSKU(soup,productURL)
        reviewCount = getReviewCount(soup)
        mainImageURL = getMainImageLinks(soup)
        starRank = getStarRank(soup)
        combine_all_size_priceString = combine_all_size_price(sizes)
        return {'size':sizes,'starRank':starRank,'reviewCount':reviewCount,'imageLink':mainImageURL,'combine_all_size_priceString':combine_all_size_priceString}
    except Exception as err:
        print('Get product detail failed:', err) 

def getMainImageLinks(soup):
    try:
        #主图
        #目前无法提取未显示的image tag，但是说不定以后可以 所以imageLinks的type先设为[]
        imageLinks = []
        #BUG-不知道如何提取hidden的itemNo-去Stack Overflow上问了
        #难怪各种方式提取 都只有一个li tag被提取出来
        imageTags = soup.find_all('li',class_=re.compile(r'itemNo'))
        for image in imageTags:
            #image.img
            #Using a tag name as an attribute will give you only the first tag by that name
            #因为li这个tag下的tags中只有一个，所以第一个就是我们需要的
            imageLinks.append(image.img['src'])
        #print("imageLinks",len(imageLinks))
        #BUG-有的imagelink怎么会有data：什么的
        return imageLinks[0]
    except Exception as err:
        print('Get Main Image Links Failed:', err) 
    
def getAll_Size_PriceForEachSKU(soup,productURL):
    try:
        #BUG-Fixed为什么(r'size_name_')而不是(r'size_name_\b+')因为是\d+不是\b+
        sizeNames = soup.find_all('li',id=re.compile(r'size_name_\d+'))
        #每个SKU的价格
        #Get all sizes
        size_price_SKU_list = [{}]
        for size_name in sizeNames:
            key = str.replace(size_name['title'],'Click to select ','')
            skuURL = 'https://www.amazon.com'+size_name['data-dp-url']
            if skuURL == 'https://www.amazon.com':
                skuURL = productURL
            #open new tab in browser to get single SKU price
            browser.get(skuURL)
            html = browser.page_source
            soup = BeautifulSoup(html, 'lxml')
            #选中价格时那这个价格就是这一页的价格 因为skuURL提取出来是""
            #如果价格是那种有折扣的话 id = "priceblock_dealprice"
            
            #Get price
            if soup.find('span',id='priceblock_ourprice'):
                price = soup.find('span',id='priceblock_ourprice').get_text() 
            elif soup.find('span',id='priceblock_dealprice'):
                price = soup.find('span',id='priceblock_dealprice').get_text()
            else:#如果缺货的话页面上就不会显示价格 find也就不会有匹配到的price
                price = "缺货"

            size_price_SKU_list.append({key:price})
        return size_price_SKU_list
    except Exception as err:
        print('Get All Size Price For Each SKU failed:', err) 

def combine_all_size_price(size_price_SKU_list):
    try:
        all_size_prices = size_price_SKU_list
        all_size_prices_combined_string = ""
        # BUG-会出现{},是哪里抓取出问题了吗
        #[{}, {'Twin': '$24.95'}, {'Twin XL': '$24.95'}, {'Full': '$25.95'}, {'Queen': '$26.95'}, {'King': '$34.95'}, {'California King': '$34.95'}] 
        #如果这个产品有size这个feature 有的产品只有颜色可选
        #TODO:考虑下如果这个产品只有颜色该怎么办
        if len(all_size_prices) != 0:
            for index,singleSKU_size_price in enumerate(all_size_prices):
                if bool(singleSKU_size_price):#如果不是空字典
                    #测试过可以将{'Twin': '$24.95'}变成['Twin']
                    key = list(singleSKU_size_price)[0]
                    value = singleSKU_size_price[key]
                    maxIndex = len(all_size_prices)-1
                    if index < maxIndex:#最后一个不用加/
                        all_size_prices_combined_string += key+":"+value+'/'
                    else:
                        all_size_prices_combined_string += key+":"+value
        return all_size_prices_combined_string
    except Exception as err:
        print("Combine failed", err)

def getStarRank(soup):
    try:
        #有两个地方会显示评分而且这两个一样 所以用find就够了
        #BUG-下面这个产品star rank的class是a-icon a-icon-star a-star-4 而不是a-icon a-icon-star a-star-4-5难怪没有get_text()这个
        #根据star rank的评分不同这个class 也会不同 2.5颗星的就变成2-5了
        #ling:https://www.amazon.com/LINENSPA-Premium-Smooth-Mattress-Protector/dp/B00MW50FKG/ref=sr_1_2_sspa?s=bedbath&ie=UTF8&qid=1522623358&sr=1-2-spons&keywords=mattress+protector&psc=1
        #starRankTag = soup.find('i',class_="a-icon a-icon-star a-star-4-5")
        #starRank = starRankTag.span.get_text()
        #Debug
        starRankTag = soup.find('span',id='acrPopover')
        starRank = starRankTag['title']
        #TODO：测试-不知道下面的会不会有速度的提升
        #修改下find div->tag1->attr
        #这样限制范围来寻找是不是会提速
        #starRankTag = soup.find('div',id='averageCustomerReviews').find('span',id='acrPopover')
        #starRankTag = starTag.span.span['title']
    except Exception as err:
        print("Get star rank failed", err)
        
    return starRank

def getReviewCount(soup):
    try:
        reviewCountTag = soup.find('span',id = 'acrCustomerReviewText')
        reviewCount = reviewCountTag.get_text()
        return reviewCount
    except Exception as err:
        print("Get Review Count failed", err)

def getAnsweredQuestionCount(soup):
    try:
        answeredQuestionCountTag = soup.find('a', id="askATFLink")
        answeredQuestionCount = answeredQuestionCountTag.span.get_text().strip()
        return answeredQuestionCount
    except Exception as err:
        print("Get Q&A failed", err)
    

def turnProductIndexToRank(product,pageNumber):
    #BUG-all rank have the same pageNumber
    #Make the soup
    html = browser.page_source
    soup = BeautifulSoup(html, 'lxml')

    #如果是九宫格
    #如：第1页第20个产品 九宫格模式 那Rank就是(1,6,2)
    if soup.find('div',class_="s-layout-picker s-grid-layout-picker"):
        #TODO:在排行那里的head增加一个九宫格排行
        productIndex = product['index']
        if productIndex <= 3:
            product['rank'] = str(pageNumber)+","+"1"+","+str(productIndex)
        elif productIndex%3 ==0:
            product['rank'] = str(pageNumber)+","+str(productIndex//3)+","+"3"
        else:
            product['rank'] = str(pageNumber)+","+str(productIndex//3 + 1)+","+str(productIndex%3)
    #四宫格-图片模式
    elif soup.find('div',class_='s-layout-picker s-image-layout-picker'):
        #Logic is the same as 九宫格 
        productIndex = product['index']
        if productIndex <= 3:
            product['rank'] = str(pageNumber)+","+"1"+","+str(productIndex)
        elif productIndex%3 ==0:
            product['rank'] = str(pageNumber)+","+str(productIndex//3)+","+"3"
        else:
            product['rank'] =  str(pageNumber)+","+str(productIndex//3 + 1)+","+str(productIndex%3)
    #剩下的就是列模式了
    #1_列模式可翻页的那种模式
    #如:https://www.amazon.com/s/ref=nb_sb_noss_2?url=search-alias%3Daps&field-keywords=tv&rh=i%3Aaps%2Ck%3Atv&ajr=0
    elif soup.find('div',class_='s-layout-picker s-list-layout-picker'):
        product['rank'] =  str(pageNumber)+','+str(productIndex)
    else:
        #2_see more的那种像厕纸一样的中间分页的那种没有翻页的那种列模式
        #如：https://www.amazon.com/gp/vs/buying-guide/sleeping-bag/459108?ie=UTF8&field-keywords=sleeping%20bag&ref_=nb_sb_ss_ime_c_1_9&url=search-alias%3Daps
        #TODO: 什么时候解决下
        #soup.find('span',id='a-autoid-0-announce').get_text() == 'See more':
        #print("See more mode")
        #product['rank'] = "See more mode"
        #Log到Excel的rank那里表示遇到了这种情况
        print("Not the normal 3 modes")
        product['rank'] = "Other mode"


#Save Rank to Excel
def saveRankToExcel(products,pageNumber,keyword):
    try:
        for product in products:
            productTitle = product['title']
            productRank = product['rank']
            wb[keyword].append([productTitle,productRank])
        wb.save("sample.xlsx")
        print('Saved')
    except Exception as err:
        print('Save Rank failed:', err)   
        wb.save("sample.xlsx")  

# 获取并储存如下信息
# ["Product Name", "Star Rank","Review Count","SKU price","Main image link","Product Link"] 
def main1():
    try:
        startTime = datetime.now()
        print("Start at:",startTime)

        global products
        # keywords to search
        #keywords = ['queen mattress protector','king mattress protector','waterproof mattress pad']
        keywords = ['mattress protector']
        for (keyword,sheetx) in zip(keywords,range(1,999)):
            global products
            # one keyword per sheet
            ws = wb.create_sheet(title=keyword)
            # initial it with one row
            ws.append(["Product Name", "Star Rank","Review Count","SKU price","Main image link","Product Link"])
            #Reset pageNumber when keyword changed
            pageNumber = 1
            search(keyword,pageNumber,sheetx,ws,targetProductNameMatching)
            #Display only the first N pages
            for pageNumber in range(2, 2):
                next_page(keyword,pageNumber,sheetx,ws,targetProductNameMatching)
            #Done getting data
            #Persist data to excel
            saveToExcel(products,keyword)
        #Process all prodcuts obtained
        #重置products 如果关键词多的话 需要重置
        #products = []
        endTime = datetime.now()
        print("Ends at:",endTime)
        #TODO:used time
    except Exception as err:
        print('出错啦', err)
        endTime = datetime.now()
        print("Ends at:",endTime)
        wb.save("sample.xlsx")
    finally:
        browser.quit()

# Title with Rank
def main():
    try:
        startTime = datetime.now()
        print("Start at:",startTime)

        global products
        # keywords to search
        #keywords = ['queen mattress protector','king mattress protector','waterproof mattress pad']
        #keywords = ['sheets']
        keywords = ['mattress pad']
        for (keyword,sheetx) in zip(keywords,range(1,999)):
            global products
            # one keyword per sheet
            ws = wb.create_sheet(title=keyword)
            ws.append(["Product Name", "Star Rank"])
            #Reset pageNumber when keyword changed
            pageNumber = 1
            search(keyword,pageNumber,sheetx,ws,targetProductNameMatching)
            #Display only the first N pages
            for pageNumber in range(2, 2):
                next_page(keyword,pageNumber,sheetx,ws,targetProductNameMatching)
            #Done getting data
            #Persist data to excel
            saveRankToExcel(products,pageNumber,keyword)
        #Process all prodcuts obtained
        #重置products 如果关键词多的话 需要重置
        #products = []
        endTime = datetime.now()
        print("Ends at:",endTime)
    except Exception as err:
        print('出错啦', err)
        endTime = datetime.now()
        print("Ends at:",endTime)
        wb.save("sample.xlsx")
    finally:
        #browser.quit()
        pass
#BUG-有的界面没有那个九宫格显示模式，怎么强制切换。
#TODO:添加一个处理总时
#TODO:保存一个条目时保存一下
#TODO:'NoneType' object has no attribute 'get_text' 处理下 排查下
if __name__ == '__main__':
    main()


#BUG-出错啦 Message: Timeout loading page after 300000ms
#BUG-Fixed不能在这里退出浏览器 不然不能搜其他的产品连接了
#browser.quit()
#注意Openpyxl添加行时.append([])要添加一个list
#.append()是添加一个单元格

#TODO:sleeping bag这种Rank如何计算
#Not the normal 3 modes
#Save Rank failed: local variable 'product' referenced before assignment
#增加进度条 不然不知道是不是卡住了
#创建Github分支不要直接在Main分支操作

#非正常的3行模式product的提取方式也不一样

#BUG-EXCEL结果和网页不一样排序