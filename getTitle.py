from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.firefox.options import Options
from selenium.webdriver import Firefox
from selenium.webdriver.common.keys import Keys
import lxml.html
#import pymongo
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import string
#Time
from datetime import datetime, date, time
import time
# Use to make a sound
import os
targetProductNameMatching = 'Maevis Bed Waterproof Mattress'

#Headless Chrome
options = webdriver.ChromeOptions()
"""options.add_argument('headless')
# Load no image makes it run faster
prefs = {"profile.managed_default_content_settings.images":2}
options.add_experimental_option("prefs",prefs)"""
options.add_argument('window-size=1200x600')
browser = webdriver.Chrome(chrome_options=options)
wait = WebDriverWait(browser, 10)
#Headless Firefox
"""options = Options()
options.add_argument('-headless')
browser = Firefox(executable_path='geckodriver', firefox_options=options)
browser.set_window_size(1200, 600)
wait = WebDriverWait(browser, 10)"""

#Excel part 
#Global variable
wb = Workbook()
products = []

def search(keyword,pageNumber):
    print('正在搜索')
    # Start
    os.system('say "Your program is start it fkuc!"')
    try:
        #美亚
        browser.get('https://www.amazon.com/')
        input = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#twotabsearchtextbox')))
        submit = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '.nav-search-submit > input:nth-child(2)')))
        input.send_keys(keyword)
        submit.click()
        get_products_title_index(keyword,pageNumber)
        # Test to compare pages
        #time.sleep(100)
    except TimeoutException:
        return search(keyword,pageNumber)


def next_page(keyword,pageNumber):
    print('正在翻页', pageNumber)
    try:
        wait.until(EC.text_to_be_present_in_element(
            (By.CSS_SELECTOR, '#pagnNextString'), 'Next Page'))
        submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#pagnNextString')))
        submit.click()
        wait.until(EC.text_to_be_present_in_element(
            (By.CSS_SELECTOR, '.pagnCur'), str(pageNumber)))
        get_products_title_index(keyword,pageNumber)
        """ 位置测试 （测试环境- Google Chrome）
        # 测试发现 除了广告位会变动外
        # 自然位的不会变动 一切匹配
        # 运行正常"""
    except TimeoutException:
        next_page(keyword,pageNumber)


def get_products_title_index(keyword,pageNumber):
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#s-results-list-atf')))
        html = browser.page_source
        soup = BeautifulSoup(html, 'lxml')
        #result_0是第一个
        content = soup.find_all(attrs={"id": re.compile(r'result_\d+')})
        print("how many result were found:",len(content))
        # 如果行数超过13行 发出警报
        if len(content) > 45:
            print('More than 15 rows.')
            os.system('say "It is more than 15 rows, you should check if it is a BUG!"')
        #如果有搜索结果
        if len(content)!=0:
            #这里除了问题 url和content数量不一致 导致url少的先结束for loop
            for index,item in enumerate(content):
                product = {
                    #是那种非产品的缺少s-access-title的，就默认给个title
                    'title': item.find(class_='s-access-title').get_text() if item.find(class_='s-access-title') else "Amazon recommendation", 
                    'index': index,#在一页里的顺位序号，每一页都会变
                    #'rank': getRank(pageNumber,index),#就算有那种AD也是准的，不影响
                }
                products.append(product)
        else:
            print("No products were found!")
    except Exception as err:
        print(err)

#获取并储存第一个广告和自然搜索的位置
#BUG-判断是否是自己的产品时出现了失误
def saveToExcel(products,pageNumber,keyword):
    try: 
        for product in products:
            productTitle = product['title']
            wb[keyword].append([productTitle])
        wb.save('Yoga titles.xlsx')
        print("Saved successfully.")
    except Exception as err:
        print('Fail to save:', err)
        wb.save('Yoga title.xlsx')

# Title with Rank
def main():
    try:
        startTime = datetime.now()
        print("Start at:",startTime)
        
        global products
        # keywords to search
        #夹棉床笠
        #keywords = ['mattress protector','queen mattress pad','mattress topper','queen mattress topper','twin mattress pad','king mattress pad','mattress cover','mattress pad cover']
        keywords = ['tpe yoga mat','yoga mat']
        #keywords = ['mattress protector','queen mattress pad']
        #There are 'fscl' 防水床笠 and 'jmcl' 夹棉床笠 这两个 type
        #whichKindOfProduct = 'fscl'
        #keywords = ['mattress protector']
        #whichKindOfProduct = 'jmcl' 
        #wb['Sheet'].cell(2,1,startTime)
        for keyword in keywords:
            wb.create_sheet(title=keyword)
            #global products
            # one keyword per sheet
            #ws = wb.create_sheet(title=keyword)
            #ws.append(["Product Name", "Star Rank"])
            #Reset pageNumber when keyword changed
            pageNumber = 1
            search(keyword,pageNumber)
            #Display only the first N pages
            for pageNumber in range(2, 3):
                next_page(keyword,pageNumber)
            #Done getting data
            #重置products 如果关键词多的话 需要重置
            saveToExcel(products,pageNumber,keyword)
            products = []
            #print("if its 0 then it's reset",len(products))
        endTime = datetime.now()
        print("Ends at:",endTime)
    except Exception as err:
        print('出错啦', err)
        endTime = datetime.now()
        print("Ends at:",endTime)
        #wb.save("sample.xlsx")
    finally:
        #browser.quit()
        pass
#BUG-有的界面没有那个九宫格显示模式，怎么强制切换。
#TODO:添加一个处理总时
#TODO:保存一个条目时保存一下
#TODO:'NoneType' object has no attribute 'get_text' 处理下 排查下
if __name__ == '__main__':
    main()


#BUG-出错啦 Message: Timeout loading page after 300000ms
#BUG-Fixed不能在这里退出浏览器 不然不能搜其他的产品连接了
#browser.quit()
#注意Openpyxl添加行时.append([])要添加一个list
#.append()是添加一个单元格

#TODO:sleeping bag这种Rank如何计算
#Not the normal 3 modes
#Save Rank failed: local variable 'product' referenced before assignment
#增加进度条 不然不知道是不是卡住了
#创建Github分支不要直接在Main分支操作

#非正常的3行模式product的提取方式也不一样

#BUG-EXCEL结果和网页不一样排序
#TODO：需要翻8页很浪费 处理下如何在找到2个最靠前的自然位后停止翻页

#