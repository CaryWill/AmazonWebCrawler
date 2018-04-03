import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import string

wb = Workbook()
wb['Sheet'].cell(1,1,"test")
wb.save('test.xlsx')