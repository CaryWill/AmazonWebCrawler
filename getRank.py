from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.firefox.options import Options
from selenium.webdriver import Firefox
from selenium.webdriver.common.keys import Keys
import lxml.html
# import pymongo
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import string
# Time
from datetime import datetime, date, time
import time
# Use to make a sound
import os
targetProductNameMatching = 'Maevis Bed Waterproof Mattress'

# Headless Chrome
options = webdriver.ChromeOptions()
# options.add_argument('headless')
# Load no image makes it run faster
prefs = {"profile.managed_default_content_settings.images":2}
# options.add_experimental_option("prefs",prefs)
# options.add_argument('window-size=1200x600')
browser = webdriver.Chrome(chrome_options=options)
wait = WebDriverWait(browser, 10)
# Headless Firefox
"""options = Options()
# options.add_argument('-headless')
browser = Firefox(executable_path='geckodriver', firefox_options=options)
browser.set_window_size(1200, 600)
wait = WebDriverWait(browser, 10)"""

# 一些全局变量
wb = Workbook()
products = []
adProducts = []
nonAdProducts = []
myproduct = []
# 亚马逊后台SKU的标题改变的话 这部分字典也要及时修改
# 词典库 用来匹配判断是否是自家产品
# 2号店 产品
# 防水床笠(fscl首字母缩写)部分
fscl = {
    'Maevis Bed Waterproof Mattress Protector Cover Pad Fitted 18 Inches Deep Pocket Premium Washable Vinyl Free - Twin XL':'TXL',
    'Waterproof Mattress Cover Protector Pad with 18 Inches Deep Pocket for Full Bed by Maevis,Full Size':'F',
    'Maevis Bed Waterproof Mattress Protector Cover Pad Fitted 18 Inches Deep Pocket Premium Washable Vinyl Free - Queen':'Q',
    'Maevis Bed Waterproof Mattress Protector Cover Pad Fitted 18 Inches Deep Pocket Premium Washable Vinyl Free - King':'K',
    'Maevis Bed Waterproof Mattress Protector Cover Pad Fitted 18 Inches Deep Pocket Premium Washable Vinyl Free - California King':'CK'
            }
# 夹棉床笠部分
jmcl = {
    'Maevis Mattress Pad Cover 100% 300TC Cotton with 8-21 Inch Deep Pocket White Overfilled Bed Mattress Topper (Down Alternative, Twin)':'T',
    'Maevis Mattress Pad Cover 100% 300TC Cotton with 8-21 Inch Deep Pocket White Overfilled Bed Mattress Topper (Down Alternative, Twin XL)':'TXL',
    'Maevis Mattress Pad Cover 100% 300TC Cotton with 8-21 Inch Deep Pocket White Overfilled Bed Mattress Topper (Down Alternative, Full)':'F',
    'Maevis Mattress Pad Cover 100% 300TC Cotton with 8-21 Inch Deep Pocket White Overfilled Bed Mattress Topper (Down Alternative, Queen)':'Q',
    'Maevis Mattress Pad Cover 100% 300TC Cotton with 8-21 Inch Deep Pocket White Overfilled Bed Mattress Topper (Down Alternative, King)':'K',
    'Maevis Mattress Pad Cover 100% 300TC Cotton with 8-21 Inch Deep Pocket White Overfilled Bed Mattress Topper (Down Alternative, California King)':'CK'
}
# 瑜伽垫
yogamat = {
    # 子SKU标题竟然都一样
    'BLC Anti-Tear TPE Yoga Mat lightweight Anti-slip 6mm Premium Exercise Mat for Yoga Fitness and GYM Workout with Carrying Strap':''
}

def search(keyword,pageNumber,productType):
    print('正在搜索',keyword)
    # Start
    # os.system('say "Your program is start now!"')
    try:
        # 美亚
        browser.get('https://www.amazon.com/')
        input = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#twotabsearchtextbox')))
        submit = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '.nav-search-submit > input:nth-child(2)')))
        input.send_keys(keyword)
        submit.click()
        get_products_title_index(keyword,pageNumber,productType)
    except TimeoutException:
        return search(keyword,pageNumber,productType)


def next_page(keyword,pageNumber,productType):
    print('正在翻页', pageNumber)
    try:
        html = browser.page_source
        soup = BeautifulSoup(html,'lxml')
        # 用来判断是否到了最后一页
        # BUG-如果在第8页结束前 就没有see more这个按钮 要提前退出 没有next page也要考虑
        # 判断是否到最后一页 （下面一个是最后一页一个不是）
        # 我发现不是最后一页的话 <span id="pagnNextString"> 这个的parent tag就是一个a tag 其中包含点击下一页后包含的下一页的链接
        # 所以可以判断如果parent tag是不是a 则到了最后一页
        # 在最后一页停止
        last_page_tag = soup.find('span',id='pagnNextString')
        #print('Type:',last_page_tag.parent.name)
        if last_page_tag.parent.name != 'a':
            print('Reach to the last page.') #未显示
            return 'Reach last page'
        else:
            #print('Not the last')
            pass
            
        if soup.find('span',id='pagnNextString'): 
            wait.until(EC.text_to_be_present_in_element(
                (By.CSS_SELECTOR, '#pagnNextString'), 'Next Page'))
            submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#pagnNextString')))
            submit.click()
            wait.until(EC.text_to_be_present_in_element(
                (By.CSS_SELECTOR, '.pagnCur'), str(pageNumber)))
        # TODO:Wants to add support for see more mode
        # BUG-why 不能用soup.find('span',class_='a-button-text') == 'See more'作为if的条件 
        elif soup.find('span',class_='a-button-text'):
            # print('See more mode!',soup.find('span',class_='a-button-text').get_text())
            submit = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="a-autoid-0"]/span/input')))
            submit.click()

        # https://www.amazon.com/s/ref=sr_pg_10?fst=p90x%3A1&rh=n%3A1055398%2Cn%3A1063252%2Cn%3A1063270%2Ck%3Amattress+game&page=10&keywords=mattress+game&ie=UTF8&qid=1523543196
        # https://www.amazon.com/s/ref=sr_pg_11?fst=p90x%3A1&rh=n%3A1055398%2Cn%3A1063252%2Cn%3A1063270%2Ck%3Amattress+game&page=11&keywords=mattress+game&ie=UTF8&qid=1523543097 
        get_products_title_index(keyword,pageNumber,productType)
        
        return 'More than 8 pages'
        
        """ 位置测试 （测试环境- Google Chrome）
        # 测试发现 除了广告位会变动外
        # 自然位的不会变动 一切匹配
        # 运行正常"""
    except TimeoutException:
        print('except')
        next_page(keyword,pageNumber,productType)

# BUG-Fixedstring indices must be integers
# 因为转换rank前product为[] 什么都没有 
def get_products_title_index(keyword,pageNumber,productType):
    try:
        html = browser.page_source
        soup = BeautifulSoup(html, 'lxml')
        # 如果需要点击的时候那么你需要等待页面显示完全 所以如果你不点击的话 完全没必要用注释里的代码
        """# 如果是三种模式中的一种 那么产品的容器可以搜s-results-list-atf
        if soup.find('ul',id='s-results-list-atf'):
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#s-results-list-atf')))
        # 如果是see more的那种
        if soup.find('ul',id='buying-guide__tabs__content'):
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR,'#buying-guide-body-v2 > div.a-section.buying-guide-btf > div.a-section.buying-guide-results__section > div > div.buying-guide-results-inner > div > div')))
        """
        content = [] # 为了兼容See more mode
        # 如果是那三种排列模式 result_0是第一个
        if soup.find(id=re.compile(r'result_\d+')):
            content = soup.find_all(attrs={"id": re.compile(r'result_\d+')})
            #print('normal')
        # 如果是see more那种模式 
        elif soup.find('li',class_='buying-guide-search-results-item'):
            #print('abnormal')
            content = soup.find_all('li',class_='s-result-item buying-guide-search-results-item')
        print("how many result were found:",len(content))
        
        # 如果行数超过13行 发出警报
        if len(content) > 45:
            print('More than 15 rows.')
            os.system('say "It is more than 15 rows, you should check if it is a BUG!"')
        
        # 如果有搜索结果
        if len(content)!=0:
            for index,item in enumerate(content):
                product = {} # 为了兼容See more mode
                # 如果是那三种排列模式 
                if soup.find(id=re.compile(r'result_\d+')): 
                    product = {
                        # 是那种非产品的缺少s-access-title的，就默认给个title
                        'title': item.find(class_='s-access-title').get_text() if item.find(class_='s-access-title') else "Amazon recommendation", 
                        'index': index+1,# 在一页里的顺位序号，每一页都会变
                        # 'rank': getRank(pageNumber,index),# 就算有那种AD也是准的，不影响
                    }
                # Sleep bag的那种see more模式
                # BUG-每次点击see more按钮 那么content会累加相当于第一页加第二页以此类推
                elif soup.find('li',class_='buying-guide-search-results-item'): 
                    product = {
                        #'title': item.find(class_='vs-carousel-title').get_text() if item.find(class_='vs-carousel-title') else "Amazon recommendation", 
                        'title': 'See more mode',
                        'index': index+1,# 在一页里的顺位序号，每一页都会变
                        # 'rank': getRank(pageNumber,index),# 就算有那种AD也是准的，不影响
                    } 
                    print('See more mode is on.')
                # 其他没见过的模式
                else:
                    product = {
                        'title': 'Other mode!', 
                        'index': index+1,# 在一页里的顺位序号，每一页都会变
                        # 'rank': getRank(pageNumber,index),# 就算有那种AD也是准的，不影响
                    } 
                    print('I do not recognize this mode, so I can not get the title from the product, please check!')
                products.append(product)
                #print(product)
                # Sort product to ad and non-ad
                identifyAndSortMyProduct(product,productType)
                # Generate Rank attr for product
                # 至少按每页转换 因为不同页可能排列模式不一样
                # TODO：这里是按每个产品来 显得有点多余（有空修改下)
                turnProductIndexToRank(product,pageNumber)
                # When to stop
                if len(adProducts)>=1 and len(nonAdProducts)>=1:
                    break 
                # print(index+1," product is processed.")
        else:
            print("No products were found!")
    except Exception as err:
        print(err)

def identifyAndSortMyProduct(product,productType):
    try:
        # 分析产品类型
        if productType == 'fscl':
            productType = fscl
        if productType == 'jmcl':
            productType = jmcl
        # Get that two err: string indices must be integers (BUG-FIXed)
        # 新增全局变量时记得这里也要修改
        if productType == 'yogamat':
            productType = yogamat
        # Sort products to ad and nonAD
        # Title may contains blank spaces
        for matchKey in list(productType):
            if matchKey in product['title'].strip(): 
                if 'Sponsored' in product['title']:
                    adProducts.append(product)
                else:
                    nonAdProducts.append(product)
                myproduct.append(product)
                break
    except Exception as err:
        print('Identify and sorting err:',err)

def getThatTwo(productType):
    try:
        if productType == 'fscl':
            productType = fscl
        if productType == 'jmcl':
            productType = jmcl
        # Get that two err: string indices must be integers (BUG-FIXed)
        # 新增全局变量时记得这里也要修改
        if productType == 'yogamat':
            productType = yogamat
        # 临时变量
        targetAdRank = ''
        targetAdAttr = ''
        targetNonAdRank = ''
        targetNonAdAttr = ''
        unifiedRankAndAttr = ''
        # 对最靠前的自然和广告位进行处理
        if len(adProducts) != 0:
            targetAdRank = adProducts[0]['rank']
            # 需要将[Sponsored]移除才能匹配到 
            targetAdAttr = productType[adProducts[0]['title'].strip().replace('[Sponsored]','')] + '广告'
        if len(nonAdProducts) != 0:
            targetNonAdRank = str(nonAdProducts[0]['rank'])
            targetNonAdAttr = productType[nonAdProducts[0]['title'].strip()] + '自然'
        # 整合 - 以 ’广告/自然‘ 为顺
        unifiedRankAndAttr = targetAdRank+'('+targetAdAttr+')'+'/'+targetNonAdRank+'('+targetNonAdAttr+')'
        # 如果在前8页搜不到自家产品则设置为默认位
        if unifiedRankAndAttr == "()/()":
            unifiedRankAndAttr = '大于8页'
        # 打印第一个广告和自然搜索的位置
        print("Two:",unifiedRankAndAttr)
        return unifiedRankAndAttr
    except Exception as err:
        print("Get that two err:",err)

# BUG-有时会出现一页超过15-20行的情况
# 按页来处理Rank
def turnProductIndexToRank(product,pageNumber):
    try:
        # Make the soup
        html = browser.page_source
        soup = BeautifulSoup(html, 'lxml')  
        productIndex = product['index']

        # 如果是九宫格和四宫格都有 默认的展示方式就可以了
        if soup.find('div',class_="s-grid-layout-picker"):
            #print("九宫格模式")
            # 选9宫格模式
            # 同时有9宫格和4宫格 
            # 都是3列 
            if soup.find('div',class_='s-image-layout-picker'):
                if productIndex <= 3:
                    product['rank'] = str(pageNumber)+"."+"1"+"."+str(productIndex)
                # 3的倍数    
                elif productIndex%3 ==0:
                    product['rank'] = str(pageNumber)+"."+str(int(productIndex/3))+"."+"3"
                else:
                    product['rank'] = str(pageNumber)+"."+str(int(productIndex//3 + 1))+"."+str(productIndex%3)
            # 剩下的就是列模式了
            # 1_列模式可翻页的那种模式
            # 如:https://www.amazon.com/s/ref=nb_sb_noss_2?url=search-alias%3Daps&field-keywords=tv&rh=i%3Aaps%2Ck%3Atv&ajr=0
            # 同时有四格和列
        elif soup.find('div',class_='s-list-layout-picker'):
            #print('可转换型列模式')
            if soup.find('div',class_='s-image-layout-picker'):
                product['rank'] =  str(pageNumber)+'.'+str(productIndex)
        # 列
        # https://www.amazon.com/s?field-keywords=sleeping+bag
        elif not soup.find('div',class_='s-list-layout-picker'):
            #print('单纯列模式-没有9宫格或四宫格按钮')
            if not soup.find('div',class_='s-image-layout-picker'):
                if not soup.find('div',class_="s-grid-layout-picker"):
                    if soup.find('span',id='pagnNextString'):
                        #os.system('say "Maybe it is a coloum mode, please check"')
                        #print("Maybe it is a coloum mode, please check,Check please!!!")
                        product['rank'] =  str(pageNumber)+'.'+str(productIndex)
                        #print("product rank:",product['rank'])
        else:
            # 2_see more的那种像厕纸一样的中间分页的那种没有翻页的那种列模式
            # 那么就没有翻页按钮 可以利用这个特点来判断
            # 如：https://www.amazon.com/gp/vs/buying-guide/sleeping-bag/459108?ie=UTF8&field-keywords=sleeping%20bag&ref_=nb_sb_ss_ime_c_1_9&url=search-alias%3Daps
            # TODO: 什么时候解决下
            # print("See more mode")
            # product['rank'] = "See more mode"
            # Log到Excel的rank那里表示遇到了这种情况
            print("Not the normal 3 modes")
            product['rank'] = "Other mode"
            # For you to check the error
            time.sleep(3000)
    except Exception as err:
        print("Convert to rank err:",err)

def getBestSellersRank(productURL):
    startTime = datetime.now()
    browser.get(productURL)
    html = browser.page_source
    soup = BeautifulSoup(html,'lxml')
    # Get all SKUs basic info
    SKUs = []
    contents = []
    bestSellerRankSequences = []
    bestSellerRankString = ''
    if soup.find('li',id=re.compile(r'size_name_\d+')):
        contents = soup.find_all('li',id=re.compile(r'size_name_\d+'))
    elif soup.find('li',id=re.compile(r'color_name_\d+')):
        contents = soup.find_all('li',id=re.compile(r'color_name_\d+')) 
    else:
        while True:
            print('????????')
    for item in contents:
        skuURL = 'https://www.amazon.com'+item['data-dp-url']
        if skuURL == 'https://www.amazon.com':
                skuURL = productURL
        sku = {
            'attr':item['title'].replace('Click to select ',''),
            'skuURL':skuURL,
        }
        SKUs.append(sku)
    # Iterate All SKUs to get detail info 
    # get best seller rank
    for sku in SKUs:
        browser.get(sku['skuURL'])
        html = browser.page_source
        soup = BeautifulSoup(html,'lxml')
        # 得到细化分类的排名
        wantedRank = ''
        # 有2个类目排名的模式 Best Sellers Rank 
        if soup.find('th',class_='prodDetSectionEntry'):
            allProdDetSectionEntries = soup.find_all('th',class_='prodDetSectionEntry')
            for entry in allProdDetSectionEntries:
                if entry.get_text().strip() == 'Best Sellers Rank':
                    # rank1
                    rank1 = entry.parent.td.span.span.get_text().strip()
                    print('rank1',rank1)
                    # rank2
                    # BUG-fixed .next_sibling 因为br的next sibling是 \n所以得call两个
                    rank2 = entry.parent.td.span.br.next_sibling.next_sibling.get_text().strip()
                    print('rank2',rank2)
                    if 'Top 100' in rank1:
                        # Get rank number in rank2
                        startIndex = rank2.find('#') 
                        # 找到第一个in 在第二个rank里面 因为rank1包含top 100 不是我们想要的
                        endIndex = rank2.find('in')
                        rankNum = rank2[startIndex:endIndex].replace('#','').strip()
                        wantedRank = rankNum
                    elif 'Top 100' in rank2:
                        # Get rank number in rank1
                        startIndex = rank1.find('#') 
                        endIndex = rank1.find('in')
                        rankNum = rank1[startIndex:endIndex].replace('#','').strip()
                        wantedRank = rankNum
                    else:
                        while True:
                            print('Two ranks do not contail Top 100')
                    
        # 只有一个类目模式的排名 Amazon Best Sellers Rank
        # BUG-gettext出现很多换行符空白
        elif soup.find('li',id='SalesRank'):
            # Top 100 在不是很细的类目下
            rank = soup.find('li',id='SalesRank').get_text()
            #print('rank',rank)
            startIndex = rank.find('#') 
            endIndex = rank.find('in')
            rankNum = rank[startIndex:endIndex].replace('#','').strip()
            wantedRank = rankNum
        else:
            while True:
                print('No rank is found')

        bestSellerRankSequences.append(wantedRank)
        sku['bestSellerRank'] = wantedRank

    for s in bestSellerRankSequences:
        bestSellerRankString += (s+'|')
    print(bestSellerRankString)
    endTime = datetime.now()
    print('Used:',endTime-startTime)
        
# Save Rank to Excel
def saveRankToExcel(keyword,keywordIndex,firstAd_N_firstNatural):
    # One keyword at a time
    try:
        # 第一个cell是（1，1） 留给日期
        # keywordIndex 是从0开始所以要+2
        wb.active.cell(1,keywordIndex+2,keyword)# keywordCell
        wb.active.cell(2,keywordIndex+2,firstAd_N_firstNatural)# rankCell
        wb.save("关键词位置统计"+".xlsx")  
        print('Saved')
    except Exception as err:
        print('Save Rank failed:', err)   
        wb.save("关键词位置统计"+str(datetime.now())+".xlsx")  
    
# Title with Rank
def main():
    try:
        # Some global products
        global products
        global nonAdProducts
        global adProducts
        # 开始
        startTime = datetime.now()
        print("Start at:",startTime)

        # ----开始----
        # 自定义部分
        # 参数部分
        #keywords = ['mattress pad','queen mattress pad','mattress topper','queen mattress topper','twin mattress pad','king mattress pad','mattress cover','mattress pad cover']
        #productType = 'jmcl'
        #keywords = ['mattress protector','waterproof mattress protector','queen mattress protector','king mattress protector','waterproof mattress pad','mattress cover']
        #keywords = ['waterproof mattress protector']
        #productType = 'fscl'
        keywords = ['tpe yoga mat','yoga mat','yoga','workout mat','fitness mat','tpe fitness yoga mat']
        #keywords = ['yoga']
        #keywords = ['tpe yoga mat']
        productType = 'yogamat'
        #keywords = ['mattress pad']
        # 表格部分-第一列
        wb.active.cell(1,1,'PC')
        wb.active.cell(2,1,str(datetime.today()))
        # -----END------

        for keywordIndex,keyword in enumerate(keywords):
            # Reset pageNumber when keyword changed
            pageNumber = 1
            search(keyword,pageNumber,productType)
            # Display only the first N pages
            for pageNumber in range(2, 8):
                # When to stop turnning page
                if len(adProducts)>=1 and len(nonAdProducts)>=1:
                    break
                else:
                    howmanypages = next_page(keyword,pageNumber,productType)
                # 搜索结果少于8页则提前停止
                print("Now is page",pageNumber)
                if howmanypages == 'Reach last page':
                    break
            # 得到最靠前的一个自然和广告位
            firstAd_N_firstNatural = getThatTwo(productType)
            # 一个关键词储存一次
            saveRankToExcel(keyword,keywordIndex,firstAd_N_firstNatural)
            # 进行搜索下一个关键词前的准备：
            # 重置一些全局变量当搜索关键词每次变得时候
            products = []
            adProducts = []
            nonAdProducts = []
        # 结束
        endTime = datetime.now()
        print("Ends at:",endTime)
        elapsed = endTime - startTime
        print("Used:",elapsed)
    except Exception as err:
        print('出错啦', err)
        endTime = datetime.now()
        print("Ends at:",endTime)
        # wb.save("关键词位置统计.xlsx")
        elapsed = endTime - startTime
        print("Used:",elapsed)
    finally:
        browser.quit()
        # Debug mode
        #pass
# BUG-有的界面没有那个九宫格显示模式，怎么强制切换。
# TODO:添加一个处理总时
# TODO:保存一个条目时保存一下
# TODO:'NoneType' object has no attribute 'get_text' 处理下 排查下
if __name__ == '__main__':
    #main()
    getBestSellersRank('https://www.amazon.com/BLC-Anti-Tear-lightweight-Anti-slip-Dark-Blue/dp/B071VVVPF3/ref=sr_1_22?ie=UTF8&qid=1523709097&sr=8-22&keywords=tpe+fitness+yoga+mat')
    

# BUG-出错啦 Message: Timeout loading page after 300000ms
# BUG-Fixed不能在这里退出浏览器 不然不能搜其他的产品连接了
# browser.quit()
# 注意Openpyxl添加行时.append([])要添加一个list
# .append()是添加一个单元格

# TODO:sleeping bag这种Rank如何计算
# Not the normal 3 modes
# 增加进度条 不然不知道是不是卡住了
# 非正常的3行模式product的提取方式也不一样