#Selenium
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver import Firefox
from selenium.webdriver.common.keys import Keys
#Beautiful Soup
from bs4 import BeautifulSoup
import lxml.html
#import pymongo
import re
#Openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
#Python
import string
import ast
from datetime import datetime, date, time
import time

#-----Done importing-------#

#Headless Chrome
options = webdriver.ChromeOptions()
#options.add_argument('headless')
# Load no image makes it run faster or you can say load web page faster 
prefs = {"profile.managed_default_content_settings.images":2}
#options.add_experimental_option("prefs",prefs)
# Windows size
options.add_argument('window-size=1200x600')
browser = webdriver.Chrome(chrome_options=options)
wait = WebDriverWait(browser, 10)
#Headless Firefox
#BUG:
#Message: Element <select class="a-native-dropdown a-button-span8" name="quantity"> is not clickable at point(792.2499847412109,243.8000030517578) because another element <span id="a-autoid-0-announce" class="a-button-text a-declarative"> obscures it
"""options = Options()
#options.add_argument('-headless')
browser = Firefox(executable_path='geckodriver', firefox_options=options)
browser.set_window_size(1200, 600)
wait = WebDriverWait(browser, 10)"""


def getStockNumber(newReleaseURL,products):
    try:
        # Program starts
        start = time.time()

        date = datetime.now()
        print('Program start at:',date)
        wb = Workbook()
        ws = wb.active
        browser.get(newReleaseURL)
        wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,'#zg_centerListWrapper')))
        html = browser.page_source
        soup = BeautifulSoup(html, 'lxml')
        content = soup.find_all('div',class_='zg_itemWrapper')
        print('how many new release were found:',len(content))
        # Done getting new release products
        for index,item in enumerate(content):
            product = {
                'title': item.img['alt'],
                'link': 'https://www.amazon.com'+item.a['href'],
                # Change string to dict - asin part
                # '{"ref":"zg_bsnr_10671048011_1","asin":"B079P3DFR4"}' is what we get first
                #'asin': ast.literal_eval(item.div['data-p13n-asin-metadata'])['asin'],
                #'mainImageURL': item.img['src']
            }
            products.append(product)
            # Reads up to 9 (10 products)
            """if index == 19: # 前10
                break"""
            # Test - Run as many times as you can to detect bugs
            """if index == 1:
                break"""
            #print(product)
        # Get stock number
        # Automatic from 1-10
        for index,product in enumerate(products):
        # index starts from 0
        # Manual 
        #for index,product in zip(range(20,21),products):
            # Manual
            #product = products[index-1]
            print('index:',index)
            # 产品详情页
            #print('link:',product['link'])
            browser.get(product['link']) 
            html = browser.page_source
            soup = BeautifulSoup(html, 'lxml')
            # Add to cart 属于产品详情页
            addToCard = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#add-to-cart-button')))
            addToCard.click()#this will update brower's current url
            # Go to my Cart
            cartURL = 'https://www.amazon.com'+soup.find('a',id='nav-cart')['href']
            #print("cart url",cartURL)
            browser.get(cartURL)
            """ After open new link you need to generate new html and make new soup
            # Update html
            html = browser.page_source
            soup = BeautifulSoup(html,'lxml')"""
            # Remember1:
            # Like element 'span' doesn't support click event
            
            # BUG-Fixed不知道如何点击数量的那个按钮
            # 使用Xpath
            # Clear the input field in order to be able to send 999
            quantity = wait.until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/div[4]/div/div[4]/div/div[2]/div[4]/form/div[2]/div/div[4]/div/div[3]/div/div/span/select/option[10]'))).click()
            quantity = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,'.a-input-text')))
            quantity.clear()
            quantity.send_keys('999')
            quantity.send_keys(Keys.RETURN)

            #BUG-Fixed不知如何得到js rendered html
            #其实js rendered html 可以配合xpath

            # 输入999后的页面跳转的反应时间
            time.sleep(3)
            # Get stock number from alert message if stock number is less than 999
            #quantityInput = browser.find_element_by_xpath('/html/body/div[1]/div[4]/div/div[4]/div/div[2]/div[4]/form/div[2]/div/div[4]/div/div[3]/div/div/input')
            quantityInput = wait.until(EC.presence_of_element_located((By.XPATH,'/html/body/div[1]/div[4]/div/div[4]/div/div[2]/div[4]/form/div[2]/div/div[4]/div/div[3]/div/div/input')))
            # If you wanna get a screenshot
            #browser.get_screenshot_as_file(str(index)+'.png') 
            inventory = quantityInput.get_attribute('value') # Type: str
            print("how many:",inventory)
            # Add inventory attr to product
            product['inventory'] = inventory
            # Alert message if have
            if inventory == '999':
                alertMessageText = "More than 999 in stock"
            else:
                alertMessage = wait.until(EC.presence_of_element_located((By.XPATH,'/html/body/div[1]/div[4]/div/div[4]/div/div[2]/div[4]/form/div[2]/div/div[4]/div[1]/div/div/div/span')))
                #alertMessage = browser.find_element_by_xpath('/html/body/div[1]/div[4]/div/div[4]/div/div[2]/div[4]/form/div[2]/div/div[4]/div[1]/div/div/div/span')
                alertMessageText = alertMessage.text
            product['inventoryAlertMessage'] = alertMessageText
            # 清除库存
            emptyCart = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,".sc-action-delete > span:nth-child(1) > input:nth-child(1)")))
            emptyCart.click()
        # Save to excel
        save(products,ws,wb,str(date)+'.xlsx')
        # Time used
        end = time.time()
        date = datetime.now()
        print("Ends at:",date)
        elapsed = end - start
        print("Used:",elapsed,'s')
    except Exception as err:
        print(err)
        save(products,ws,wb,str(date)+'.xlsx')
        # Time used
        end = time.time()
        date = datetime.now()
        print("Ends at:",date)
        elapsed = end - start
        print("Used:",elapsed)
    finally:
        # Don't forget quitting your browser
        browser.quit()

def save(products,ws,wb,wbName):
    # First row (Concise one with only inventory)
    ws.append(['Date','Order','Title','Inventory','Alert Message']) 
    # Fisrt row (Detailed one)
    #ws.append(['Date','Order','Title','Star Rank','Review Count','QNA Count','Main Image Link','SKU Link','Inventory','Alert Message'])
    # If KeyError: 'inventory' happens
    # Run it again will do
    # Auto
    for index,product in enumerate(products):
    # Manual    
    #for index,product in zip(range(20,21),products):
        # Manual 
        #product = products[index-1]
        # If you want more of the product
        #productInfo = getProductDetail(product['link'])
        # Just title and inventory
        #ws.append([datetime.now(),index,product['title'],int(product['inventory'])])
        # added inventory alert message
        ws.append([datetime.now(),index,product['title'],product['inventory'],product['inventoryAlertMessage']])
        # Default String format Version
        #ws.append([datetime.now(),index,product['title'],productInfo['starRank'],productInfo['reviewCount'],productInfo['QNA'],productInfo['imageLink'],product['link'],product['inventory'],product['inventoryAlertMessage']])
        # Number format Version
        #ws.append([datetime.now(),index,product['title'],float(productInfo['starRank']),float(productInfo['reviewCount']),float(productInfo['QNA']),productInfo['imageLink'],product['link'],int(product['inventory']),product['inventoryAlertMessage']])
    wb.save(wbName)
    browser.quit()

def main():
    try:
        products = []
        #newReleaseURL = 'https://www.amazon.com/gp/new-releases/home-garden/10671048011/ref=zg_bs_tab_t_bsnr'
        newReleaseURL = 'https://www.amazon.com/gp/new-releases/home-garden/3732781/ref=zg_bs_tab_t_bsnr'
        getStockNumber(newReleaseURL,products)
    except Exception as err:
        print('Err on main:',err)
    finally:
        browser.quit()
    
if __name__ == '__main__':
    main()